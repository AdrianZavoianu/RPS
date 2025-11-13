"""
Export service for RPS projects.

Provides export functionality for result data in Excel and CSV formats.
"""

from typing import Optional, Callable
from pathlib import Path
from dataclasses import dataclass
from datetime import datetime
import json
import pandas as pd

from config.result_config import RESULT_CONFIGS
from database.models import GlobalResultsCache


@dataclass
class ExportOptions:
    """Options for exporting result data (MVP: single result type)."""

    result_set_id: int
    result_type: str  # e.g., "Drifts_X", "QuadRotations"
    output_path: Path
    format: str = "excel"  # "excel" or "csv"
    include_maxmin: bool = False  # For future implementation


@dataclass
class ProjectExportExcelOptions:
    """Options for Excel project export."""
    output_path: Path
    include_all_results: bool = True
    result_set_ids: Optional[list[int]] = None  # None = all sets


class ExportService:
    """Service for exporting result data.

    MVP Implementation: Exports single result type to Excel or CSV.

    Args:
        project_context: ProjectContext instance with session access
        result_service: ResultDataService instance for data retrieval
    """

    APP_VERSION = "2.7"

    def __init__(self, project_context, result_service):
        """Initialize export service.

        Args:
            project_context: ProjectContext with session() and session_factory()
            result_service: ResultDataService for retrieving datasets
        """
        self.context = project_context
        self.result_service = result_service

    def export_result_type(
        self,
        options: ExportOptions,
        progress_callback: Optional[Callable[[str, int, int], None]] = None
    ) -> None:
        """Export single result type to file.

        Args:
            options: Export options specifying what and where to export
            progress_callback: Optional callback(message, current, total) for progress updates

        Raises:
            ValueError: If result_type is unknown
            IOError: If file cannot be written
        """
        # Step 1: Validate result type and get configuration
        config = RESULT_CONFIGS.get(options.result_type)
        if not config:
            raise ValueError(f"Unknown result type: {options.result_type}")

        # Step 2: Extract direction and base type (handles directionless results)
        direction = self._extract_direction(options.result_type, config)
        base_type = self._extract_base_type(options.result_type)

        if progress_callback:
            progress_callback("Loading data...", 1, 3)

        # Step 3: Get dataset using correct ResultDataService API (no element_id parameter!)
        dataset = self.result_service.get_standard_dataset(
            result_type=base_type,
            direction=direction,
            result_set_id=options.result_set_id
        )

        if progress_callback:
            progress_callback("Writing file...", 2, 3)

        # Step 4: Write to file based on format
        if options.format == "excel":
            dataset.data.to_excel(options.output_path, index=False, engine='openpyxl')
        else:  # csv
            dataset.data.to_csv(options.output_path, index=False)

        if progress_callback:
            progress_callback("Export complete!", 3, 3)

    def _extract_direction(self, result_type: str, config) -> str:
        """Extract direction from result type name.

        Handles directionless results (QuadRotations, MinAxial) by returning empty string.

        Args:
            result_type: Full result type name (e.g., "Drifts_X", "QuadRotations")
            config: ResultConfig for this result type

        Returns:
            Direction string (e.g., "X", "V2") or empty string for directionless

        Examples:
            "Drifts_X" → "X"
            "WallShears_V2" → "V2"
            "QuadRotations" → ""
            "MinAxial" → ""
        """
        # Directionless results have empty direction_suffix
        if not config.direction_suffix:
            return ""

        # Extract direction from result_type name (last part after underscore)
        if "_" in result_type:
            return result_type.split("_")[-1]

        # Fallback to config's direction suffix
        return config.direction_suffix

    def _extract_base_type(self, result_type: str) -> str:
        """Extract base result type from full result type name.

        Args:
            result_type: Full result type name (e.g., "Drifts_X", "QuadRotations")

        Returns:
            Base type name without direction

        Examples:
            "Drifts_X" → "Drifts"
            "WallShears_V2" → "WallShears"
            "QuadRotations" → "QuadRotations"
        """
        # Split on last underscore to separate direction
        if "_" in result_type:
            return result_type.rsplit("_", 1)[0]

        # No underscore means no direction (already base type)
        return result_type

    def get_available_result_types(self, result_set_id: int) -> list[str]:
        """Get list of available result types for a result set.

        Queries the cache to determine which result types have data.

        Args:
            result_set_id: ID of result set to query

        Returns:
            Sorted list of result type names with data
        """
        available = []

        # Use context.session() to create session (correct pattern)
        with self.context.session() as session:
            # Query GlobalResultsCache for available types
            global_types = session.query(GlobalResultsCache.result_type).filter(
                GlobalResultsCache.result_set_id == result_set_id
            ).distinct().all()

            available.extend([rt[0] for rt in global_types])

        return sorted(set(available))

    def build_filename(self, result_type: str, format: str) -> str:
        """Build filename matching Old_scripts naming convention.

        Args:
            result_type: Full result type name (e.g., "Drifts_X")
            format: Output format ("excel" or "csv")

        Returns:
            Filename with appropriate extension

        Examples:
            build_filename("Drifts_X", "excel") → "Drifts_X.xlsx"
            build_filename("QuadRotations", "csv") → "QuadRotations.csv"
        """
        ext = "xlsx" if format == "excel" else "csv"
        return f"{result_type}.{ext}"

    def _get_result_config(self, result_type: str):
        """Get result config for a result type.

        Args:
            result_type: Full result type name (e.g., "Drifts_X")

        Returns:
            ResultTypeConfig instance

        Raises:
            ValueError: If result type is unknown
        """
        config = RESULT_CONFIGS.get(result_type)
        if not config:
            raise ValueError(f"Unknown result type: {result_type}")
        return config

    def get_element_export_dataframe(self, result_type: str, result_set_id: int):
        """Get combined DataFrame for all elements of a result type.

        Args:
            result_type: Full result type (e.g., "WallShears_V2", "QuadRotations")
            result_set_id: Result set ID

        Returns:
            pd.DataFrame with all elements combined, or None if no data
        """
        import pandas as pd
        from database.models import ElementResultsCache, Element, Story
        from database.repository import ElementCacheRepository

        # Get all elements with data for this result type
        with self.context.session() as session:
            element_cache_repo = ElementCacheRepository(session)

            # Query all cache entries for this result type
            entries = session.query(
                ElementResultsCache, Element, Story
            ).join(
                Element, ElementResultsCache.element_id == Element.id
            ).join(
                Story, ElementResultsCache.story_id == Story.id
            ).filter(
                ElementResultsCache.result_set_id == result_set_id,
                ElementResultsCache.result_type == result_type
            ).order_by(
                Element.name,  # Group by element
                ElementResultsCache.story_sort_order.desc()  # Stories bottom to top
            ).all()

            if not entries:
                return None

            # Build rows
            rows = []
            for cache_entry, element, story in entries:
                row = {"Element": element.name, "Story": story.name}
                # Merge results_matrix (load case columns)
                if cache_entry.results_matrix:
                    row.update(cache_entry.results_matrix)
                rows.append(row)

            df = pd.DataFrame(rows)
            return df

    # ===== PROJECT EXPORT TO EXCEL =====

    def export_project_excel(
        self,
        options: ProjectExportExcelOptions,
        progress_callback: Optional[Callable[[str, int, int], None]] = None
    ) -> None:
        """Export complete project to Excel workbook with metadata.

        Creates .xlsx file with:
        - README sheet (human-readable overview)
        - Metadata sheets (result sets, load cases, stories, elements)
        - Result data sheets (one per result type)
        - IMPORT_DATA sheet (hidden JSON for re-import)

        Args:
            options: Export options
            progress_callback: Optional callback(message, current, total)
        """
        from openpyxl import load_workbook
        from openpyxl.styles import Font

        total_steps = 10

        # Step 1: Gather metadata
        if progress_callback:
            progress_callback("Gathering project metadata...", 1, total_steps)

        metadata = self._gather_project_metadata()

        # Step 2: Create Excel writer
        if progress_callback:
            progress_callback("Creating Excel workbook...", 2, total_steps)

        with pd.ExcelWriter(options.output_path, engine='openpyxl') as writer:
            # Step 3: Write README sheet
            if progress_callback:
                progress_callback("Writing README sheet...", 3, total_steps)
            self._write_readme_sheet(writer, metadata)

            # Step 4: Write metadata sheets
            if progress_callback:
                progress_callback("Writing metadata sheets...", 4, total_steps)
            self._write_metadata_sheets(writer, metadata)

            # Step 5-8: Write result data sheets
            if progress_callback:
                progress_callback("Writing result data sheets...", 5, total_steps)

            result_sheets = self._write_result_data_sheets(
                writer, metadata, options,
                lambda msg, curr, tot: progress_callback(msg, 5 + curr, total_steps) if progress_callback else None
            )

            # Step 9: Write IMPORT_DATA sheet
            if progress_callback:
                progress_callback("Writing import metadata...", 9, total_steps)
            self._write_import_data_sheet(writer, metadata, result_sheets)

        # Step 10: Apply formatting (reopen with openpyxl)
        if progress_callback:
            progress_callback("Applying formatting...", 10, total_steps)
        self._apply_excel_formatting(options.output_path)

        if progress_callback:
            progress_callback("Export complete!", total_steps, total_steps)

    def _gather_project_metadata(self) -> dict:
        """Gather all project metadata for export."""
        from database.catalog_base import get_catalog_session
        from database.catalog_repository import CatalogProjectRepository
        from services.project_service import get_project_summary
        from database.models import ResultSet, ResultCategory, LoadCase, Story, Element

        # Get catalog entry
        catalog_session = get_catalog_session()
        try:
            catalog_repo = CatalogProjectRepository(catalog_session)
            catalog_project = catalog_repo.get_by_slug(self.context.slug)
            if not catalog_project:
                raise ValueError(f"Project '{self.context.slug}' not found in catalog")
        finally:
            catalog_session.close()

        # Get project summary
        summary = get_project_summary(self.context)

        # Query project database
        with self.context.session() as session:
            # Result sets
            result_sets = session.query(ResultSet).filter(
                ResultSet.project_id == self.result_service.project_id
            ).all()

            # Result categories (no project_id, linked via result_set_id)
            result_set_ids = [rs.id for rs in result_sets]
            result_categories = session.query(ResultCategory).filter(
                ResultCategory.result_set_id.in_(result_set_ids)
            ).all() if result_set_ids else []

            # Load cases
            load_cases = session.query(LoadCase).filter(
                LoadCase.project_id == self.result_service.project_id
            ).all()

            # Stories
            stories = session.query(Story).filter(
                Story.project_id == self.result_service.project_id
            ).order_by(Story.sort_order).all()

            # Elements
            elements = session.query(Element).filter(
                Element.project_id == self.result_service.project_id
            ).order_by(Element.element_type, Element.name).all()

            return {
                'catalog_project': catalog_project,
                'summary': summary,
                'result_sets': result_sets,
                'result_categories': result_categories,
                'load_cases': load_cases,
                'stories': stories,
                'elements': elements,
            }

    def _write_readme_sheet(self, writer, metadata: dict) -> None:
        """Write README sheet with project overview."""
        catalog = metadata['catalog_project']
        summary = metadata['summary']

        # Build README content
        readme_lines = [
            ["PROJECT INFORMATION"],
            ["==================="],
            [""],
            ["Project Name:", catalog.name],
            ["Description:", catalog.description or ""],
            ["Created:", catalog.created_at.strftime("%Y-%m-%d %H:%M:%S")],
            ["Exported:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["RPS Version:", self.APP_VERSION],
            [""],
            ["DATABASE SUMMARY"],
            ["================"],
            [""],
            ["Result Sets:", summary.result_sets],
            ["Load Cases:", summary.load_cases],
            ["Stories:", summary.stories],
            ["Elements:", len(metadata['elements'])],
            [""],
            ["IMPORT INSTRUCTIONS"],
            ["==================="],
            [""],
            ["To import this project into RPS:"],
            ["1. Open RPS application"],
            ["2. Click 'Import Project' on Projects page"],
            ["3. Select this Excel file"],
            ["4. Review project summary"],
            ["5. Click 'Import'"],
            [""],
            ["The hidden 'IMPORT_DATA' sheet contains metadata required for re-import."],
            ["Do NOT delete or modify this sheet."],
        ]

        df = pd.DataFrame(readme_lines)
        df.to_excel(writer, sheet_name="README", index=False, header=False)

    def _write_metadata_sheets(self, writer, metadata: dict) -> None:
        """Write metadata sheets (Result Sets, Load Cases, Stories, Elements)."""
        # Result Sets sheet
        result_sets_data = [
            {
                "Name": rs.name,
                "Description": rs.description or "",
                "Created At": rs.created_at.strftime("%Y-%m-%d %H:%M:%S") if rs.created_at else "",
            }
            for rs in metadata['result_sets']
        ]
        if result_sets_data:
            pd.DataFrame(result_sets_data).to_excel(writer, sheet_name="Result Sets", index=False)

        # Load Cases sheet
        load_cases_data = [
            {
                "Name": lc.name,
                "Description": lc.description or "",
            }
            for lc in metadata['load_cases']
        ]
        if load_cases_data:
            pd.DataFrame(load_cases_data).to_excel(writer, sheet_name="Load Cases", index=False)

        # Stories sheet
        stories_data = [
            {
                "Name": s.name,
                "Sort Order": s.sort_order,
                "Elevation": s.elevation or 0.0,
            }
            for s in metadata['stories']
        ]
        if stories_data:
            pd.DataFrame(stories_data).to_excel(writer, sheet_name="Stories", index=False)

        # Elements sheet
        elements_data = [
            {
                "Name": e.name,
                "Unique Name": e.unique_name or "",
                "Element Type": e.element_type,
            }
            for e in metadata['elements']
        ]
        if elements_data:
            pd.DataFrame(elements_data).to_excel(writer, sheet_name="Elements", index=False)

    def _write_result_data_sheets(
        self,
        writer,
        metadata: dict,
        options: ProjectExportExcelOptions,
        progress_callback: Optional[Callable] = None
    ) -> dict:
        """Write result data sheets (one per result type).

        Returns:
            Dict mapping result type to sheet name for IMPORT_DATA
        """
        from database.models import ElementResultsCache

        result_sheets = {'global': [], 'element': []}

        # Determine which result sets to export
        if options.result_set_ids:
            result_sets = [rs for rs in metadata['result_sets'] if rs.id in options.result_set_ids]
        else:
            result_sets = metadata['result_sets']

        # For simplicity, export first result set only
        if not result_sets:
            return result_sheets

        result_set = result_sets[0]

        # Discover available result types from NORMALIZED tables (source of truth)
        with self.context.session() as session:
            from database.models import StoryDrift, StoryAcceleration, StoryForce, StoryDisplacement

            # Export Drifts (read from story_drifts table)
            directions = session.query(StoryDrift.direction).distinct().all()
            for direction, in directions:
                config_key = f"Drifts_{direction}"
                df = self._get_normalized_drift_dataframe(session, result_set.id, direction)
                if df is not None and not df.empty:
                    sheet_name = config_key[:31]
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    result_sheets['global'].append(config_key)
                    if progress_callback:
                        progress_callback(f"Exported {config_key}", 0, 1)

            # Export Accelerations (read from story_accelerations table)
            directions = session.query(StoryAcceleration.direction).distinct().all()
            for direction, in directions:
                config_key = f"Accelerations_{direction}"
                df = self._get_normalized_acceleration_dataframe(session, result_set.id, direction)
                if df is not None and not df.empty:
                    sheet_name = config_key[:31]
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    result_sheets['global'].append(config_key)
                    if progress_callback:
                        progress_callback(f"Exported {config_key}", 0, 1)

            # Export Forces (read from story_forces table)
            directions = session.query(StoryForce.direction).distinct().all()
            for direction, in directions:
                config_key = f"Forces_{direction}"
                df = self._get_normalized_force_dataframe(session, result_set.id, direction)
                if df is not None and not df.empty:
                    sheet_name = config_key[:31]
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    result_sheets['global'].append(config_key)
                    if progress_callback:
                        progress_callback(f"Exported {config_key}", 0, 1)

            # Export Displacements (read from story_displacements table)
            directions = session.query(StoryDisplacement.direction).distinct().all()
            for direction, in directions:
                config_key = f"Displacements_{direction}"
                df = self._get_normalized_displacement_dataframe(session, result_set.id, direction)
                if df is not None and not df.empty:
                    sheet_name = config_key[:31]
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    result_sheets['global'].append(config_key)
                    if progress_callback:
                        progress_callback(f"Exported {config_key}", 0, 1)

            # Element results (still use cache for elements)
            element_types = session.query(ElementResultsCache.result_type).filter(
                ElementResultsCache.result_set_id == result_set.id
            ).distinct().all()

            for result_type, in element_types:
                # Get combined element data
                df = self.get_element_export_dataframe(result_type, result_set.id)

                if df is not None and not df.empty:
                    sheet_name = result_type[:31]
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    result_sheets['element'].append(result_type)

                    if progress_callback:
                        progress_callback(f"Exported {result_type}", 0, 1)

        return result_sheets

    def _get_normalized_drift_dataframe(self, session, result_set_id: int, direction: str) -> pd.DataFrame:
        """Get drift data from story_drifts table."""
        from database.models import StoryDrift, Story, LoadCase

        query = session.query(
            Story.name.label('Story'),
            LoadCase.name.label('LoadCase'),
            StoryDrift.drift
        ).join(
            Story, StoryDrift.story_id == Story.id
        ).join(
            LoadCase, StoryDrift.load_case_id == LoadCase.id
        ).filter(
            StoryDrift.direction == direction
        ).order_by(
            StoryDrift.story_sort_order,
            LoadCase.name
        )

        results = query.all()
        if not results:
            return None

        # Pivot to wide format: Story | LC1 | LC2 | ...
        data = {}
        for story, load_case, value in results:
            if story not in data:
                data[story] = {'Story': story}
            data[story][load_case] = value

        return pd.DataFrame(list(data.values()))

    def _get_normalized_acceleration_dataframe(self, session, result_set_id: int, direction: str) -> pd.DataFrame:
        """Get acceleration data from story_accelerations table."""
        from database.models import StoryAcceleration, Story, LoadCase

        query = session.query(
            Story.name.label('Story'),
            LoadCase.name.label('LoadCase'),
            StoryAcceleration.acceleration
        ).join(
            Story, StoryAcceleration.story_id == Story.id
        ).join(
            LoadCase, StoryAcceleration.load_case_id == LoadCase.id
        ).filter(
            StoryAcceleration.direction == direction
        ).order_by(
            StoryAcceleration.story_sort_order,
            LoadCase.name
        )

        results = query.all()
        if not results:
            return None

        # Pivot to wide format
        data = {}
        for story, load_case, value in results:
            if story not in data:
                data[story] = {'Story': story}
            data[story][load_case] = value

        return pd.DataFrame(list(data.values()))

    def _get_normalized_force_dataframe(self, session, result_set_id: int, direction: str) -> pd.DataFrame:
        """Get force data from story_forces table."""
        from database.models import StoryForce, Story, LoadCase

        query = session.query(
            Story.name.label('Story'),
            LoadCase.name.label('LoadCase'),
            StoryForce.force
        ).join(
            Story, StoryForce.story_id == Story.id
        ).join(
            LoadCase, StoryForce.load_case_id == LoadCase.id
        ).filter(
            StoryForce.direction == direction
        ).order_by(
            StoryForce.story_sort_order,
            LoadCase.name
        )

        results = query.all()
        if not results:
            return None

        # Pivot to wide format
        data = {}
        for story, load_case, value in results:
            if story not in data:
                data[story] = {'Story': story}
            data[story][load_case] = value

        return pd.DataFrame(list(data.values()))

    def _get_normalized_displacement_dataframe(self, session, result_set_id: int, direction: str) -> pd.DataFrame:
        """Get displacement data from story_displacements table."""
        from database.models import StoryDisplacement, Story, LoadCase

        query = session.query(
            Story.name.label('Story'),
            LoadCase.name.label('LoadCase'),
            StoryDisplacement.displacement
        ).join(
            Story, StoryDisplacement.story_id == Story.id
        ).join(
            LoadCase, StoryDisplacement.load_case_id == LoadCase.id
        ).filter(
            StoryDisplacement.direction == direction
        ).order_by(
            StoryDisplacement.story_sort_order,
            LoadCase.name
        )

        results = query.all()
        if not results:
            return None

        # Pivot to wide format
        data = {}
        for story, load_case, value in results:
            if story not in data:
                data[story] = {'Story': story}
            data[story][load_case] = value

        return pd.DataFrame(list(data.values()))

    def _write_import_data_sheet(self, writer, metadata: dict, result_sheets: dict) -> None:
        """Write IMPORT_DATA sheet with complete database dump."""
        with self.context.session() as session:
            from database.models import (
                StoryDrift, StoryAcceleration, StoryForce, StoryDisplacement,
                GlobalResultsCache, ElementResultsCache
            )

            # Build complete import data with ALL per-project database tables
            import_data = {
                "version": self.APP_VERSION,
                "export_timestamp": datetime.now().isoformat(),
                "project": {
                    "name": metadata['catalog_project'].name,
                    "slug": metadata['catalog_project'].slug,
                    "description": metadata['catalog_project'].description or "",
                    "created_at": metadata['catalog_project'].created_at.isoformat(),
                },
                "result_sets": [
                    {
                        "name": rs.name,
                        "description": rs.description or "",
                        "created_at": rs.created_at.isoformat() if rs.created_at else None,
                    }
                    for rs in metadata['result_sets']
                ],
                "result_categories": [
                    {
                        "category_name": rc.category_name,
                        "result_set_name": next((rs.name for rs in metadata['result_sets'] if rs.id == rc.result_set_id), None),
                        "category_type": rc.category_type,
                    }
                    for rc in metadata.get('result_categories', [])
                ],
                "load_cases": [
                    {"name": lc.name, "description": lc.description or ""}
                    for lc in metadata['load_cases']
                ],
                "stories": [
                    {"name": s.name, "sort_order": s.sort_order, "elevation": s.elevation or 0.0}
                    for s in metadata['stories']
                ],
                "elements": [
                    {"name": e.name, "unique_name": e.unique_name or "", "element_type": e.element_type}
                    for e in metadata['elements']
                ],
                "result_sheet_mapping": result_sheets,

                # Complete normalized table dumps
                "normalized_data": {
                    "story_drifts": self._serialize_story_drifts(session),
                    "story_accelerations": self._serialize_story_accelerations(session),
                    "story_forces": self._serialize_story_forces(session),
                    "story_displacements": self._serialize_story_displacements(session),
                    "absolute_maxmin_drifts": self._serialize_absolute_maxmin_drifts(session),
                    "quad_rotations": self._serialize_quad_rotations(session),
                    "wall_shears": self._serialize_wall_shears(session),
                },

                # Complete cache table dumps
                "cache_data": {
                    "global_results_cache": self._serialize_global_cache(session),
                    "element_results_cache": self._serialize_element_cache(session),
                }
            }

            # Write as compact JSON, split into chunks to avoid Excel cell limit (32,767 chars)
            json_str = json.dumps(import_data, separators=(',', ':'))

            # Split into chunks of 30,000 characters
            chunk_size = 30000
            chunks = [json_str[i:i+chunk_size] for i in range(0, len(json_str), chunk_size)]

            # Write each chunk as a row
            df = pd.DataFrame(chunks, columns=["import_metadata"])
            df.to_excel(writer, sheet_name="IMPORT_DATA", index=False)

    def _serialize_story_drifts(self, session) -> list:
        """Serialize all story_drifts table data."""
        from database.models import StoryDrift, Story, LoadCase

        results = session.query(
            Story.name.label('story_name'),
            LoadCase.name.label('load_case_name'),
            StoryDrift.direction,
            StoryDrift.drift,
            StoryDrift.max_drift,
            StoryDrift.min_drift,
            StoryDrift.story_sort_order
        ).join(Story, StoryDrift.story_id == Story.id
        ).join(LoadCase, StoryDrift.load_case_id == LoadCase.id
        ).all()

        return [
            {
                "story_name": r.story_name,
                "load_case_name": r.load_case_name,
                "direction": r.direction,
                "drift": r.drift,
                "max_drift": r.max_drift,
                "min_drift": r.min_drift,
                "story_sort_order": r.story_sort_order
            }
            for r in results
        ]

    def _serialize_story_accelerations(self, session) -> list:
        """Serialize all story_accelerations table data."""
        from database.models import StoryAcceleration, Story, LoadCase, ResultCategory, ResultSet

        results = session.query(
            Story.name.label('story_name'),
            LoadCase.name.label('load_case_name'),
            ResultSet.name.label('result_set_name'),
            ResultCategory.category_name.label('result_category_name'),
            StoryAcceleration.direction,
            StoryAcceleration.acceleration,
            StoryAcceleration.max_acceleration,
            StoryAcceleration.min_acceleration,
            StoryAcceleration.story_sort_order
        ).join(Story, StoryAcceleration.story_id == Story.id
        ).join(LoadCase, StoryAcceleration.load_case_id == LoadCase.id
        ).join(ResultCategory, StoryAcceleration.result_category_id == ResultCategory.id
        ).join(ResultSet, ResultCategory.result_set_id == ResultSet.id
        ).all()

        return [
            {
                "story_name": r.story_name,
                "load_case_name": r.load_case_name,
                "result_set_name": r.result_set_name,
                "result_category_name": r.result_category_name,
                "direction": r.direction,
                "acceleration": r.acceleration,
                "max_acceleration": r.max_acceleration,
                "min_acceleration": r.min_acceleration,
                "story_sort_order": r.story_sort_order
            }
            for r in results
        ]

    def _serialize_story_forces(self, session) -> list:
        """Serialize all story_forces table data."""
        from database.models import StoryForce, Story, LoadCase, ResultCategory, ResultSet

        results = session.query(
            Story.name.label('story_name'),
            LoadCase.name.label('load_case_name'),
            ResultSet.name.label('result_set_name'),
            ResultCategory.category_name.label('result_category_name'),
            StoryForce.direction,
            StoryForce.location,
            StoryForce.force,
            StoryForce.max_force,
            StoryForce.min_force,
            StoryForce.story_sort_order
        ).join(Story, StoryForce.story_id == Story.id
        ).join(LoadCase, StoryForce.load_case_id == LoadCase.id
        ).join(ResultCategory, StoryForce.result_category_id == ResultCategory.id
        ).join(ResultSet, ResultCategory.result_set_id == ResultSet.id
        ).all()

        return [
            {
                "story_name": r.story_name,
                "load_case_name": r.load_case_name,
                "result_set_name": r.result_set_name,
                "result_category_name": r.result_category_name,
                "direction": r.direction,
                "location": r.location,
                "force": r.force,
                "max_force": r.max_force,
                "min_force": r.min_force,
                "story_sort_order": r.story_sort_order
            }
            for r in results
        ]

    def _serialize_story_displacements(self, session) -> list:
        """Serialize all story_displacements table data."""
        from database.models import StoryDisplacement, Story, LoadCase, ResultCategory, ResultSet

        results = session.query(
            Story.name.label('story_name'),
            LoadCase.name.label('load_case_name'),
            ResultSet.name.label('result_set_name'),
            ResultCategory.category_name.label('result_category_name'),
            StoryDisplacement.direction,
            StoryDisplacement.displacement,
            StoryDisplacement.max_displacement,
            StoryDisplacement.min_displacement,
            StoryDisplacement.story_sort_order
        ).join(Story, StoryDisplacement.story_id == Story.id
        ).join(LoadCase, StoryDisplacement.load_case_id == LoadCase.id
        ).join(ResultCategory, StoryDisplacement.result_category_id == ResultCategory.id
        ).join(ResultSet, ResultCategory.result_set_id == ResultSet.id
        ).all()

        return [
            {
                "story_name": r.story_name,
                "load_case_name": r.load_case_name,
                "result_set_name": r.result_set_name,
                "result_category_name": r.result_category_name,
                "direction": r.direction,
                "displacement": r.displacement,
                "max_displacement": r.max_displacement,
                "min_displacement": r.min_displacement,
                "story_sort_order": r.story_sort_order
            }
            for r in results
        ]

    def _serialize_absolute_maxmin_drifts(self, session) -> list:
        """Serialize all absolute_maxmin_drifts table data."""
        from database.models import AbsoluteMaxMinDrift, Story, LoadCase, ResultSet

        results = session.query(
            ResultSet.name.label('result_set_name'),
            Story.name.label('story_name'),
            Story.sort_order.label('story_sort_order'),
            LoadCase.name.label('load_case_name'),
            AbsoluteMaxMinDrift.direction,
            AbsoluteMaxMinDrift.absolute_max_drift,
            AbsoluteMaxMinDrift.sign,
            AbsoluteMaxMinDrift.original_max,
            AbsoluteMaxMinDrift.original_min
        ).join(ResultSet, AbsoluteMaxMinDrift.result_set_id == ResultSet.id
        ).join(Story, AbsoluteMaxMinDrift.story_id == Story.id
        ).join(LoadCase, AbsoluteMaxMinDrift.load_case_id == LoadCase.id
        ).all()

        return [
            {
                "result_set_name": r.result_set_name,
                "story_name": r.story_name,
                "story_sort_order": r.story_sort_order,
                "load_case_name": r.load_case_name,
                "direction": r.direction,
                "absolute_max_drift": r.absolute_max_drift,
                "sign": r.sign,
                "original_max": r.original_max,
                "original_min": r.original_min
            }
            for r in results
        ]

    def _serialize_quad_rotations(self, session) -> list:
        """Serialize all quad_rotations table data."""
        from database.models import QuadRotation, Story, LoadCase, Element

        results = session.query(
            Element.name.label('element_name'),
            Story.name.label('story_name'),
            LoadCase.name.label('load_case_name'),
            QuadRotation.rotation,
            QuadRotation.max_rotation,
            QuadRotation.min_rotation,
            QuadRotation.story_sort_order
        ).join(Element, QuadRotation.element_id == Element.id
        ).join(Story, QuadRotation.story_id == Story.id
        ).join(LoadCase, QuadRotation.load_case_id == LoadCase.id
        ).all()

        return [
            {
                "element_name": r.element_name,
                "story_name": r.story_name,
                "load_case_name": r.load_case_name,
                "rotation": r.rotation,
                "max_rotation": r.max_rotation,
                "min_rotation": r.min_rotation,
                "story_sort_order": r.story_sort_order
            }
            for r in results
        ]

    def _serialize_wall_shears(self, session) -> list:
        """Serialize all wall_shears table data."""
        from database.models import WallShear, Story, LoadCase, Element

        results = session.query(
            Element.name.label('element_name'),
            Story.name.label('story_name'),
            LoadCase.name.label('load_case_name'),
            WallShear.direction,
            WallShear.location,
            WallShear.force,
            WallShear.max_force,
            WallShear.min_force,
            WallShear.story_sort_order
        ).join(Element, WallShear.element_id == Element.id
        ).join(Story, WallShear.story_id == Story.id
        ).join(LoadCase, WallShear.load_case_id == LoadCase.id
        ).all()

        return [
            {
                "element_name": r.element_name,
                "story_name": r.story_name,
                "load_case_name": r.load_case_name,
                "direction": r.direction,
                "location": r.location,
                "force": r.force,
                "max_force": r.max_force,
                "min_force": r.min_force,
                "story_sort_order": r.story_sort_order
            }
            for r in results
        ]

    def _serialize_global_cache(self, session) -> list:
        """Serialize all global_results_cache table data."""
        from database.models import GlobalResultsCache, Story, ResultSet

        results = session.query(
            ResultSet.name.label('result_set_name'),
            Story.name.label('story_name'),
            GlobalResultsCache.result_type,
            GlobalResultsCache.story_sort_order,
            GlobalResultsCache.results_matrix
        ).join(ResultSet, GlobalResultsCache.result_set_id == ResultSet.id
        ).join(Story, GlobalResultsCache.story_id == Story.id
        ).all()

        return [
            {
                "result_set_name": r.result_set_name,
                "story_name": r.story_name,
                "result_type": r.result_type,
                "story_sort_order": r.story_sort_order,
                "results_matrix": r.results_matrix
            }
            for r in results
        ]

    def _serialize_element_cache(self, session) -> list:
        """Serialize all element_results_cache table data."""
        from database.models import ElementResultsCache, Story, ResultSet, Element

        results = session.query(
            ResultSet.name.label('result_set_name'),
            Element.name.label('element_name'),
            Story.name.label('story_name'),
            ElementResultsCache.result_type,
            ElementResultsCache.story_sort_order,
            ElementResultsCache.results_matrix
        ).join(ResultSet, ElementResultsCache.result_set_id == ResultSet.id
        ).join(Element, ElementResultsCache.element_id == Element.id
        ).join(Story, ElementResultsCache.story_id == Story.id
        ).all()

        return [
            {
                "result_set_name": r.result_set_name,
                "element_name": r.element_name,
                "story_name": r.story_name,
                "result_type": r.result_type,
                "story_sort_order": r.story_sort_order,
                "results_matrix": r.results_matrix
            }
            for r in results
        ]

    def _apply_excel_formatting(self, file_path: Path) -> None:
        """Apply formatting to Excel workbook (bold headers, hide IMPORT_DATA sheet)."""
        from openpyxl import load_workbook
        from openpyxl.styles import Font

        wb = load_workbook(file_path)

        # Format README sheet
        if "README" in wb.sheetnames:
            ws = wb["README"]
            for row in ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.font = Font(bold=True, size=14)

        # Hide IMPORT_DATA sheet
        if "IMPORT_DATA" in wb.sheetnames:
            wb["IMPORT_DATA"].sheet_state = "hidden"

        wb.save(file_path)
