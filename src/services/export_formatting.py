"""Excel formatting helpers for exports."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font


def apply_excel_formatting(file_path: Path) -> None:
    """Apply formatting to Excel workbook (bold headers, hide IMPORT_DATA sheet)."""
    wb = load_workbook(file_path)

    if "README" in wb.sheetnames:
        ws = wb["README"]
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True, size=14)

    if "IMPORT_DATA" in wb.sheetnames:
        wb["IMPORT_DATA"].sheet_state = "hidden"

    wb.save(file_path)
